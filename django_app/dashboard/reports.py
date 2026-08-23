"""Формирование отчётов: Excel (openpyxl) и CSV."""
import csv
import io

from django.utils import timezone

# Пароль защиты книги/листов от форматирования (защита от подделки отчётов)
REPORT_PROTECT_PASSWORD = 'molvest-report'


def _xlsx_style():
    """Общие стили Excel-отчётов: заголовок шапки таблицы (без заливки —
    фон ячеек в отчётах не используется), шрифт заголовка листа."""
    from openpyxl.styles import Font
    return (
        Font(bold=True),
        Font(bold=True, size=14),
    )


def _table_border():
    """Тонкая окантовка таблицы отчёта."""
    from openpyxl.styles import Border, Side
    thin = Side(style='thin', color='FF7F7F7F')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _finalize_sheet(ws):
    """Финальное оформление листа отчёта: без сетки ячеек на фоне и с защитой
    от форматирования — ячейки нельзя править, заливать, менять границы и
    шрифты (нельзя подделать отчёт). Разрешён только просмотр и выделение.
    """
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    ws.protection.password = REPORT_PROTECT_PASSWORD
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.insertHyperlinks = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.pivotTables = False
    ws.protection.objects = False
    ws.protection.scenarios = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def _protect_workbook(wb):
    """Защита структуры книги: нельзя добавлять/удалять/переименовывать листы."""
    wb.security.lockStructure = True
    wb.security.lockWindows = False
    wb.security.workbookPassword = REPORT_PROTECT_PASSWORD


# ---------------------------------------------------------------------------
# Простои
# ---------------------------------------------------------------------------

def build_downtime_xlsx(meta, events):
    """Отчёт о простоях: события с длительностью, сводка по линиям.

    Без сетки и заливок на фоне; таблица — только с окантовкой; листы и книга
    защищены от форматирования (нельзя подделать отчёт).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    header_font, title_font = _xlsx_style()
    border = _table_border()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Простои'
    ws.append(['Отчёт о простоях'])
    ws['A1'].font = title_font
    ws.append([f'Период: {meta["from_str"]} — {meta["to_str"]}'])
    ws.append([f'Сформирован: {meta["generated_at"]}'])
    ws.append([])
    headers = ['№', 'Линия', 'Цех', 'Продукт', 'Начало простоя', 'Окончание', 'Длительность, мин', 'Статус']
    ws.append(headers)
    for cell in ws[5]:
        cell.font = header_font
        cell.border = border

    per_line = {}
    for i, e in enumerate(events, start=1):
        row_idx = ws.max_row + 1
        ws.append([
            i, e['line_name'], e['shop_name'],
            f"{e['product_code']} — {e['product_name']}",
            timezone.localtime(e['start']).strftime('%d.%m.%Y %H:%M'),
            timezone.localtime(e['end']).strftime('%d.%m.%Y %H:%M'),
            e['minutes'],
            'продолжается' if e['ongoing'] else 'завершён',
        ])
        for cell in ws[row_idx]:
            cell.border = border
        per_line[e['line_name']] = per_line.get(e['line_name'], 0) + e['minutes']

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=6, value='ИТОГО, мин:').font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=meta['total_minutes']).font = Font(bold=True)
    for cell in ws[total_row]:
        cell.border = border

    for col, w in zip('ABCDEFGH', [5, 42, 26, 40, 18, 18, 16, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A6'
    _finalize_sheet(ws)

    if per_line:
        ws2 = wb.create_sheet('Сводка по линиям')
        ws2.append(['Линия', 'Простоев', 'Минут'])
        for cell in ws2[1]:
            cell.font = header_font
            cell.border = border
        for name, minutes in per_line.items():
            row_idx = ws2.max_row + 1
            ws2.append([name, sum(1 for e in events if e['line_name'] == name), minutes])
            for cell in ws2[row_idx]:
                cell.border = border
        ws2.column_dimensions['A'].width = 42
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 10
        _finalize_sheet(ws2)

    _protect_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return meta['filename_xlsx'], buf.getvalue()


def build_downtime_csv(meta, events):
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow(['Отчёт о простоях'])
    writer.writerow([f'Период: {meta["from_str"]} — {meta["to_str"]}'])
    writer.writerow([f'Сформирован: {meta["generated_at"]}'])
    writer.writerow([])
    writer.writerow(['№', 'Линия', 'Цех', 'Продукт', 'Начало простоя', 'Окончание', 'Длительность, мин', 'Статус'])
    for i, e in enumerate(events, start=1):
        writer.writerow([
            i, e['line_name'], e['shop_name'],
            f"{e['product_code']} — {e['product_name']}",
            timezone.localtime(e['start']).strftime('%d.%m.%Y %H:%M'),
            timezone.localtime(e['end']).strftime('%d.%m.%Y %H:%M'),
            e['minutes'],
            'продолжается' if e['ongoing'] else 'завершён',
        ])
    writer.writerow(['ИТОГО, мин:', '', '', '', '', '', meta['total_minutes'], ''])
    return meta['filename_csv'], buf.getvalue().encode('utf-8-sig')


# ---------------------------------------------------------------------------
# Универсальный экспорт таблиц отчётов (новый движок отчётов)
# ---------------------------------------------------------------------------

def _fmt_cell(v):
    if isinstance(v, (int, float)):
        return v
    return str(v) if v is not None else ''


def _xlsx_hex(color):
    """Цвет '#RRGGBB' -> 'RRGGBB' (openpyxl ждёт hex без решётки).

    Принимает также 'RRGGBB' и 'AARRGGBB' (8 символов). Возвращает None,
    если значение не похоже на цвет.
    """
    if not color:
        return None
    s = str(color).strip().lstrip('#')
    if len(s) not in (6, 8) or not all(ch in '0123456789abcdefABCDEF' for ch in s):
        return None
    return s.upper()


def _write_report_sheet(ws, meta, table, write_meta=True):
    """Заполняет лист Excel одной таблицей отчёта.

    Формат: объединённый заголовок отчёта по ширине листа, «Сформирован:»,
    шапка отчёта (мета-строки, каждая объединена по ширине), строка по центру
    (если есть), шапка таблицы (жирная, по центру; для колонок «Код продукта |
    Заводской код | …» — двухстрочная: «Код:» объединено по 2 колонкам +
    «Продукта | Заводской»), данные, итог, примечание. Сетка ячеек на фоне и
    заливки не используются — таблица только с окантовкой. Лист защищён от
    форматирования.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    header_font, _ = _xlsx_style()
    border = _table_border()

    def merge_row(row_idx, ncols):
        if ncols > 1:
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=ncols)

    columns = table.get('columns') or []
    ncols = max(len(columns), 1)

    # 1) Заголовок отчёта — объединён по ширине, по центру
    ws.append([meta.get('title', 'Отчёт')])
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws.row_dimensions[1].height = 22.5
    merge_row(1, ncols)

    # 2) Сформирован
    ws.append([f'Сформирован: {meta.get("generated_at", "")}'])
    ws[ws.max_row][0].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    merge_row(ws.max_row, ncols)

    # 3) Шапка отчёта (мета-строки)
    if write_meta and meta.get('report_meta'):
        for label, value in meta['report_meta']:
            ws.append([f'{label} {value}'])
            cell = ws[ws.max_row][0]
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            merge_row(ws.max_row, ncols)

    # 4) Строка по центру (например «Смена 1»)
    if table.get('title_row'):
        ws.append([table['title_row']])
        ws[ws.max_row][0].font = Font(bold=True, size=13)
        ws[ws.max_row][0].alignment = Alignment(horizontal='center')
        merge_row(ws.max_row, ncols)

    # 6) Шапка таблицы — жирная по центру, без заливки, с окантовкой
    header_row = ws.max_row
    if columns:
        two_row = (len(columns) >= 2
                   and str(columns[0]).startswith('Код')
                   and str(columns[1]).startswith('Заводской'))
        if two_row:
            ws.append(['Код:', ''] + list(columns[2:]))
            ws.merge_cells(start_row=ws.max_row, start_column=1,
                           end_row=ws.max_row, end_column=2)
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.append(['Продукта', 'Заводской'] + [''] * max(0, len(columns) - 2))
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            ws.append(columns)
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 7) Данные (с окантовкой)
    for row in table.get('rows') or []:
        row_idx = ws.max_row + 1
        ws.append([_fmt_cell(v) for v in row])
        for cell in ws[row_idx]:
            cell.border = border

    # 8) Итог — жирный, с окантовкой
    if table.get('total_row'):
        row_idx = ws.max_row + 1
        ws.append([_fmt_cell(v) for v in table['total_row']])
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.border = border

    if table.get('note'):
        ws.append([])
        ws.append([table['note']])

    if columns:
        # Ширина колонок по количеству символов в заголовке, данных и итоге
        col_lens = [len(str(c)) for c in columns]
        for row in table.get('rows') or []:
            for i, cell in enumerate(row):
                if i < len(col_lens):
                    col_lens[i] = max(col_lens[i], len(str(cell)))
        tr = table.get('total_row')
        if tr:
            for i, cell in enumerate(tr):
                if i < len(col_lens):
                    col_lens[i] = max(col_lens[i], len(str(cell)))
        caps = []
        for ln in col_lens:
            width = min(70, max(8, int(ln * 1.2) + 2))
            caps.append(width)
        for col_idx, width in enumerate(caps, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Страховка: если текст длиннее ширины колонки — перенос на новую строку
        for row_idx in range(1, ws.max_row + 1):
            for col_idx, cw in enumerate(caps, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and len(str(cell.value)) > cw:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.freeze_panes = f'A{header_row + 1}'
    _finalize_sheet(ws)


def export_tables_xlsx(meta, tables):
    """Каждая таблица отчёта — отдельный лист Excel (мета-строки — на первом).

    Листы без сетки и заливок, таблица с окантовкой, книга защищена от
    форматирования."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for idx, table in enumerate(tables, start=1):
        ws = wb.create_sheet(f'Лист{idx}')
        _write_report_sheet(ws, meta, table, write_meta=(idx == 1))
    _protect_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return meta['filename_xlsx'], buf.getvalue()


def _write_day_chart_sheet(wb, ws, meta, chart_data, data_sheet_name='Данные'):
    """Заполняет лист графиком продукции за сутки.

    Сутки делятся на 6 частей по 4 часа; каждая часть — отдельная диаграмма,
    где каждый столбец — одна минута (240 столбцов на диаграмму). Цвет
    столбца = цвет продукта за эту минуту, над столбцом мелким шрифтом (8 pt)
    выводится количество продукции. Все 6 диаграмм помещаются на одном листе
    A4 (альбомная ориентация, вписываются в одну страницу). Внизу листа —
    легенда: код, цвет и название всех продуктов, попавших на графики.
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.properties import PageSetupProperties

    labels = chart_data.get('labels') or []                       # 'HH:MM' по минутам
    data = (chart_data.get('datasets') or [{}])[0].get('data') or []
    details = chart_data.get('details') or []

    def parse_minute(lbl):
        """'HH:MM' -> минута суток (0..1439); None, если не похоже на время."""
        try:
            hh, mm = str(lbl).split(':')
            return int(hh) * 60 + int(mm)
        except (ValueError, AttributeError):
            return None

    # Минута суток -> (количество, детали продукта). Порядок не важен:
    # дальше всегда обращаемся по индексу минуты.
    by_minute = {}
    for lbl, val, det in zip(labels, data, details):
        m = parse_minute(lbl)
        if m is None or m < 0 or m >= 1440:
            continue
        by_minute[m] = (int(val or 0), det)

    def hour_label(h):
        """'24:00' -> '00:00' (час за полночью — это ноль часов новых суток)."""
        return f'{h % 24:02d}:00'

    def minute_label(m):
        return f'{m // 60:02d}:{m % 60:02d}'

    # 6 секций по 4 часа (минуты 00:00..23:59)
    sections = [(0, 4), (4, 8), (8, 12), (12, 16), (16, 20), (20, 24)]

    # Продукты на графиках (порядок появления): код -> {name, color}
    products = {}
    order = []
    for m in range(1440):
        det = by_minute.get(m, (0, None))[1]
        code = det.get('code') if det else None
        if not code or code in products:
            continue
        products[code] = {'name': det.get('name'), 'color': det.get('color')}
        order.append(code)

    # Страница: A4, альбомная ориентация, вписать в одну страницу
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    ws['A1'] = meta.get('title', 'График за сутки')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Период: {meta.get("period_label", "")} · Счетчик: {meta.get("counter", "")}'
    ws['A2'].font = Font(bold=True, size=11)

    # Данные графика — на отдельном скрытом листе (пользователю не видны)
    ws_data = wb.create_sheet(data_sheet_name)
    ws_data.sheet_state = 'hidden'
    _finalize_sheet(ws_data)

    # Данные секций на скрытом листе: минута/значение в колонках A..L (строки 1-240)
    for si, (h0, h1) in enumerate(sections):
        col = 1 + si * 2  # A, C, E, G, I, K
        row = 1
        for m in range(h0 * 60, h1 * 60):
            ws_data.cell(row=row, column=col, value=minute_label(m))
            ws_data.cell(row=row, column=col + 1, value=by_minute.get(m, (0, None))[0])
            row += 1

    # 6 диаграмм: 2 колонки x 3 ряда (вписываются на один лист A4).
    # Размер и расстановка — как в образце: 15 x 7.5 см, колонки A/L,
    # ряды 4/21/38 (диаграммы не накладываются друг на друга).
    anchors = [
        ('A4', 0), ('L4', 1),
        ('A21', 2), ('L21', 3),
        ('A38', 4), ('L38', 5),
    ]
    for (anchor, si) in anchors:
        h0, h1 = sections[si]
        col = 1 + si * 2
        chart = BarChart()
        chart.type = 'col'
        chart.title = f'{hour_label(h0)} – {hour_label(h1)}'
        chart.width = 15.0
        chart.height = 7.5
        data_ref = Reference(ws_data, min_col=col + 1, min_row=1, max_row=240)
        cats_ref = Reference(ws_data, min_col=col, min_row=1, max_row=240)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        # Ось категорий — внизу (openpyxl по умолчанию ставит её слева; в
        # образце Excel она внизу), подписи времени — каждые 5 минут.
        chart.x_axis.delete = False
        chart.x_axis.axPos = 'b'
        chart.x_axis.tickLblSkip = 5
        chart.x_axis.tickMarkSkip = 5
        chart.y_axis.delete = False
        chart.y_axis.axPos = 'l'
        # Подписи количества над каждым столбиком — мелким шрифтом (7 pt)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.numFmt = '0'
        chart.dataLabels.dLblPos = 'outEnd'
        tx_pr = RichText()
        tx_pr.p = [Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=700)))]
        chart.dataLabels.txPr = tx_pr
        # Цвет каждого столбика = цвет продукта за эту минуту
        series = chart.series[0]
        for j, m in enumerate(range(h0 * 60, h1 * 60)):
            dp = DataPoint(idx=j)
            det = by_minute.get(m, (0, None))[1]
            dp.graphicalProperties.solidFill = _xlsx_hex(det.get('color') if det else None) or '6C757D'
            series.data_points.append(dp)
        chart.legend = None
        ws.add_chart(chart, anchor)

    # Легенда внизу листа: код, цвет (образец цвета), название продукта.
    # Расположена ниже последнего ряда диаграмм (как в образце). Без заливки
    # фона: только окантовка; цвет продукта — заливкой ячейки-образца.
    if order:
        border = _table_border()
        start_row = 55
        ws.cell(row=start_row, column=1, value='Продукты на графиках:')
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        header_row = start_row + 1
        for ci, h in enumerate(['Код', 'Цвет', 'Название продукта']):
            cell = ws.cell(row=header_row, column=ci + 1, value=h)
            cell.font = Font(bold=True)
            cell.border = border
        for i, code in enumerate(order, start=1):
            p = products[code]
            r = header_row + i
            code_cell = ws.cell(row=r, column=1, value=code)
            code_cell.border = border
            color_cell = ws.cell(row=r, column=2)
            hexc = _xlsx_hex(p.get('color'))
            if hexc:
                color_cell.fill = PatternFill('solid', fgColor=hexc)
            color_cell.alignment = Alignment(horizontal='center')
            color_cell.border = border
            name_cell = ws.cell(row=r, column=3, value=p.get('name') or '')
            name_cell.border = border
        for col, w in zip('ABC', [10, 12, 50]):
            ws.column_dimensions[col].width = w

    _finalize_sheet(ws)


def build_day_chart_xlsx(meta, chart_data):
    """График продукции за сутки в Excel: 6 диаграмм по 4 часа на одном
    листе A4 (альбомная ориентация), каждый столбец — минута, цвет столбца
    = цвет продукта, подписи количества над столбцами мелким шрифтом,
    внизу — легенда продуктов. Лист защищён от форматирования."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'График за сутки'
    _write_day_chart_sheet(wb, ws, meta, chart_data, data_sheet_name='Данные')
    _protect_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return meta.get('filename_xlsx', 'chart_day.xlsx'), buf.getvalue()


def export_tables_csv(meta, tables):
    """Каждая таблица отчёта — блок строк в одном CSV."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow([meta.get('title', 'Отчёт')])
    writer.writerow([f'Сформирован: {meta.get("generated_at", "")}'])
    # Шапка отчёта (мета-строки)
    if meta.get('report_meta'):
        for label, value in meta['report_meta']:
            writer.writerow([f'{label} {value}'])
    writer.writerow([])
    for table in tables:
        if table.get('title_row'):
            writer.writerow([table['title_row']])
        columns = table.get('columns') or []
        if columns:
            writer.writerow(columns)
        for row in table.get('rows') or []:
            writer.writerow([_fmt_cell(v) for v in row])
        if table.get('total_row'):
            writer.writerow([_fmt_cell(v) for v in table['total_row']])
        writer.writerow([])
    return meta['filename_csv'], buf.getvalue().encode('utf-8-sig')


# ---------------------------------------------------------------------------
# Пакетный экспорт нескольких отчётов (сравнение) — один файл, лист на отчёт
# ---------------------------------------------------------------------------

def export_reports_bundle_xlsx(items):
    """Несколько отчётов — один Excel-файл, отдельный лист на каждый отчёт.

    items: список отчётов:
      {'kind': 'table', 'meta': {...}, 'tables': [...]} — лист с заголовком,
        «Сформирован», мета-строками и всеми таблицами отчёта (пустая строка
        между таблицами);
      {'kind': 'chart', 'meta': {...}, 'chart': {...}} — лист с графиком за
        сутки (6 диаграмм по 4 часа на листе A4).

    Все листы — без сетки ячеек и заливок фона (таблица только с окантовкой),
    книга защищена от форматирования (нельзя подделать отчёт).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    header_font, _ = _xlsx_style()
    border = _table_border()

    wb = Workbook()
    wb.remove(wb.active)
    for idx, item in enumerate(items, start=1):
        ws = wb.create_sheet(f'Отчет {idx}')
        if item.get('kind') == 'chart':
            _write_day_chart_sheet(wb, ws, item.get('meta') or {}, item.get('chart') or {},
                                   data_sheet_name=f'Данные {idx}')
            continue
        meta = item.get('meta') or {}
        tables = item.get('tables') or []
        ncols = max([len((t.get('columns') or [])) for t in tables] or [1])

        def merge_row(row_idx):
            if ncols > 1:
                ws.merge_cells(start_row=row_idx, start_column=1,
                               end_row=row_idx, end_column=ncols)

        # Заголовок отчёта + «Сформирован» + мета-строки — в начале листа
        ws.append([meta.get('title', 'Отчёт')])
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        ws.row_dimensions[1].height = 22.5
        merge_row(1)

        ws.append([f'Сформирован: {meta.get("generated_at", "")}'])
        ws[ws.max_row][0].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        merge_row(ws.max_row)

        if meta.get('report_meta'):
            for label, value in meta['report_meta']:
                ws.append([f'{label} {value}'])
                cell = ws[ws.max_row][0]
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                merge_row(ws.max_row)

        header_row = None
        for table in tables:
            if table.get('title_row'):
                ws.append([table['title_row']])
                ws[ws.max_row][0].font = Font(bold=True, size=13)
                ws[ws.max_row][0].alignment = Alignment(horizontal='center')
                merge_row(ws.max_row)
            columns = table.get('columns') or []
            if columns:
                two_row = (len(columns) >= 2
                           and str(columns[0]).startswith('Код')
                           and str(columns[1]).startswith('Заводской'))
                if two_row:
                    ws.append(['Код:', ''] + list(columns[2:]))
                    ws.merge_cells(start_row=ws.max_row, start_column=1,
                                   end_row=ws.max_row, end_column=2)
                    if header_row is None:
                        header_row = ws.max_row
                    for cell in ws[ws.max_row]:
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.append(['Продукта', 'Заводской'] + [''] * max(0, len(columns) - 2))
                    if header_row is None:
                        header_row = ws.max_row
                    for cell in ws[ws.max_row]:
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    ws.append(columns)
                    if header_row is None:
                        header_row = ws.max_row
                    for cell in ws[ws.max_row]:
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            for row in table.get('rows') or []:
                row_idx = ws.max_row + 1
                ws.append([_fmt_cell(v) for v in row])
                for cell in ws[row_idx]:
                    cell.border = border
            if table.get('total_row'):
                row_idx = ws.max_row + 1
                ws.append([_fmt_cell(v) for v in table['total_row']])
                for cell in ws[row_idx]:
                    cell.font = Font(bold=True)
                    cell.border = border
            if table.get('note'):
                ws.append([])
                ws.append([table['note']])
            ws.append([])  # разделитель между таблицами

        # Ширина колонок — по символам (максимум по всем таблицам листа)
        caps = {}
        for table in tables:
            columns = table.get('columns') or []
            if not columns:
                continue
            col_lens = [len(str(c)) for c in columns]
            for row in table.get('rows') or []:
                for i, cell in enumerate(row):
                    if i < len(col_lens):
                        col_lens[i] = max(col_lens[i], len(str(cell)))
            tr = table.get('total_row')
            if tr:
                for i, cell in enumerate(tr):
                    if i < len(col_lens):
                        col_lens[i] = max(col_lens[i], len(str(cell)))
            for i, ln in enumerate(col_lens, start=1):
                caps[i] = max(caps.get(i, 0), min(70, max(8, int(ln * 1.2) + 2)))
        for col_idx, width in caps.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        if header_row:
            ws.freeze_panes = f'A{header_row + 1}'
        _finalize_sheet(ws)

    _protect_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return 'reports_comparison.xlsx', buf.getvalue()


def export_reports_bundle_csv(items):
    """Несколько отчётов — один CSV: блоки отчётов разделены пустой строкой."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    for item in items:
        meta = item.get('meta') or {}
        writer.writerow([meta.get('title', 'Отчёт')])
        writer.writerow([f'Сформирован: {meta.get("generated_at", "")}'])
        if meta.get('report_meta'):
            for label, value in meta['report_meta']:
                writer.writerow([f'{label} {value}'])
        writer.writerow([])
        for table in item.get('tables') or []:
            if table.get('title_row'):
                writer.writerow([table['title_row']])
            columns = table.get('columns') or []
            if columns:
                writer.writerow(columns)
            for row in table.get('rows') or []:
                writer.writerow([_fmt_cell(v) for v in row])
            if table.get('total_row'):
                writer.writerow([_fmt_cell(v) for v in table['total_row']])
            writer.writerow([])
        writer.writerow([])
    return 'reports_comparison.csv', buf.getvalue().encode('utf-8-sig')
